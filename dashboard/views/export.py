"""Excel export page."""

from __future__ import annotations

import io
from datetime import date, timedelta

import streamlit as st

from core.database import Database
from core.models import VALID_PIPELINE_STAGES


def render() -> None:
    """Render the export page."""
    st.header("Export Leads")

    db: Database = st.session_state["db"]

    # ── Filter controls ─────────────────────────────────────────
    st.subheader("Export Filters")

    col1, col2 = st.columns(2)
    with col1:
        stages = st.multiselect(
            "Pipeline Stages",
            options=sorted(VALID_PIPELINE_STAGES),
            default=sorted(VALID_PIPELINE_STAGES),
            key="export_stages",
        )
        score_threshold = st.slider(
            "Minimum Score",
            min_value=0,
            max_value=100,
            value=0,
            key="export_min_score",
        )
    with col2:
        date_range = st.date_input(
            "Discovery Date Range",
            value=(date.today() - timedelta(days=365), date.today()),
            key="export_date_range",
        )

    # ── Gather leads ────────────────────────────────────────────
    all_leads = []
    for stage in stages:
        leads_in_stage = db.get_all_leads(stage=stage, min_score=score_threshold, limit=5000)
        all_leads.extend(leads_in_stage)

    # Apply date filter
    if len(date_range) == 2:
        start_date, end_date = date_range
        all_leads = [
            l for l in all_leads
            if start_date <= l.discovery_date <= end_date
        ]

    # De-duplicate (in case any overlap)
    seen_ids: set[int] = set()
    unique_leads = []
    for lead in all_leads:
        if lead.id not in seen_ids:
            seen_ids.add(lead.id)
            unique_leads.append(lead)
    all_leads = unique_leads

    # Sort by score descending
    all_leads.sort(key=lambda l: l.score, reverse=True)

    st.info(f"Found **{len(all_leads)}** leads matching the filters.")

    # ── Preview table ───────────────────────────────────────────
    if all_leads:
        st.subheader("Preview")
        preview_rows = []
        for lead in all_leads:
            preview_rows.append({
                "ID": lead.id,
                "Company": lead.company_name,
                "Contact": lead.contact_name or "",
                "Title": lead.contact_title or "",
                "Email": lead.email or "",
                "Phone": lead.phone or "",
                "Score": lead.score,
                "Type": lead.company_type.replace("_", " ").title(),
                "Stage": lead.pipeline_stage,
                "Address": lead.address or "",
                "Building": lead.building_type or "",
                "SqFt": lead.sqft or "",
                "Tenants": lead.num_tenants or "",
                "IT Provider": lead.current_it_provider or "",
                "Discovery": lead.discovery_date.isoformat(),
            })

        st.dataframe(preview_rows, use_container_width=True, hide_index=True, height=300)

    # ── Export button ───────────────────────────────────────────
    st.divider()
    if all_leads:
        xlsx_bytes = _build_xlsx(all_leads)
        st.download_button(
            label="Download Excel (.xlsx)",
            data=xlsx_bytes,
            file_name="salesleads_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    else:
        st.warning("No leads to export. Adjust the filters above.")


def _build_xlsx(leads: list) -> bytes:
    """Build an .xlsx file in memory and return the bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")

    headers = [
        "ID", "Company", "Contact Name", "Contact Title", "Email", "Phone",
        "LinkedIn", "Website", "Address", "Building Type", "SqFt", "Tenants",
        "IT Provider", "Tech Signals", "Recent News", "Sources", "Score",
        "Stage", "Company Type", "Discovery Date", "Notes",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, lead in enumerate(leads, 2):
        ws.cell(row=row_num, column=1, value=lead.id)
        ws.cell(row=row_num, column=2, value=lead.company_name)
        ws.cell(row=row_num, column=3, value=lead.contact_name)
        ws.cell(row=row_num, column=4, value=lead.contact_title)
        ws.cell(row=row_num, column=5, value=lead.email or "")
        ws.cell(row=row_num, column=6, value=lead.phone or "")
        ws.cell(row=row_num, column=7, value=lead.linkedin_url or "")
        ws.cell(row=row_num, column=8, value=lead.website or "")
        ws.cell(row=row_num, column=9, value=lead.address)
        ws.cell(row=row_num, column=10, value=lead.building_type)
        ws.cell(row=row_num, column=11, value=lead.sqft or "")
        ws.cell(row=row_num, column=12, value=lead.num_tenants or "")
        ws.cell(row=row_num, column=13, value=lead.current_it_provider or "")
        ws.cell(row=row_num, column=14, value=", ".join(lead.tech_signals))
        ws.cell(row=row_num, column=15, value=", ".join(lead.recent_news))
        ws.cell(row=row_num, column=16, value=", ".join(lead.sources))
        ws.cell(row=row_num, column=17, value=lead.score)
        ws.cell(row=row_num, column=18, value=lead.pipeline_stage)
        ws.cell(row=row_num, column=19, value=lead.company_type)
        ws.cell(row=row_num, column=20, value=lead.discovery_date.isoformat())
        ws.cell(row=row_num, column=21, value=lead.qualification_notes)

    # Auto-width (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
