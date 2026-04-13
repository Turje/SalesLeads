"""Excel export route."""

from __future__ import annotations

import io

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from api.deps import get_db
from api.schemas import ExportRequest
from core.models import EnrichedLead, VALID_PIPELINE_STAGES

router = APIRouter(prefix="/api/export", tags=["export"])


def _build_xlsx(leads: list[EnrichedLead]) -> bytes:
    """Build an .xlsx file in memory and return the raw bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")

    headers = [
        "ID",
        "Company",
        "Company Type",
        "Contact Name",
        "Contact Title",
        "Email",
        "Phone",
        "LinkedIn",
        "Website",
        "Address",
        "Borough",
        "Neighborhood",
        "Building Type",
        "SqFt",
        "Tenants",
        "Year Built",
        "Floors",
        "Employees",
        "Building ISP",
        "Available ISPs",
        "Equipment",
        "Building Summary",
        "IT Provider",
        "Tech Signals",
        "Recent News",
        "Sources",
        "Score",
        "Stage",
        "Discovery Date",
        "Notes",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, lead in enumerate(leads, 2):
        ws.cell(row=row_num, column=1, value=lead.id)
        ws.cell(row=row_num, column=2, value=lead.company_name)
        ws.cell(row=row_num, column=3, value=lead.company_type)
        ws.cell(row=row_num, column=4, value=lead.contact_name)
        ws.cell(row=row_num, column=5, value=lead.contact_title)
        ws.cell(row=row_num, column=6, value=lead.email or "")
        ws.cell(row=row_num, column=7, value=lead.phone or "")
        ws.cell(row=row_num, column=8, value=lead.linkedin_url or "")
        ws.cell(row=row_num, column=9, value=lead.website or "")
        ws.cell(row=row_num, column=10, value=lead.address)
        ws.cell(row=row_num, column=11, value=lead.borough)
        ws.cell(row=row_num, column=12, value=lead.neighborhood)
        ws.cell(row=row_num, column=13, value=lead.building_type)
        ws.cell(row=row_num, column=14, value=lead.sqft or "")
        ws.cell(row=row_num, column=15, value=lead.num_tenants or "")
        ws.cell(row=row_num, column=16, value=lead.year_built or "")
        ws.cell(row=row_num, column=17, value=lead.floors or "")
        ws.cell(row=row_num, column=18, value=lead.num_employees or "")
        ws.cell(row=row_num, column=19, value=lead.building_isp or "")
        ws.cell(row=row_num, column=20, value=", ".join(lead.available_isps))
        ws.cell(row=row_num, column=21, value=str(lead.equipment) if lead.equipment else "")
        ws.cell(row=row_num, column=22, value=lead.building_summary)
        ws.cell(row=row_num, column=23, value=lead.current_it_provider or "")
        ws.cell(row=row_num, column=24, value=", ".join(lead.tech_signals))
        ws.cell(row=row_num, column=25, value=", ".join(lead.recent_news))
        ws.cell(row=row_num, column=26, value=", ".join(lead.sources))
        ws.cell(row=row_num, column=27, value=lead.score)
        ws.cell(row=row_num, column=28, value=lead.pipeline_stage)
        ws.cell(row=row_num, column=29, value=lead.discovery_date.isoformat())
        ws.cell(row=row_num, column=30, value=lead.qualification_notes)

    # Auto-width (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.post("/xlsx")
def export_xlsx(body: ExportRequest) -> StreamingResponse:
    db = get_db()

    # Determine which stages to export
    stages = body.stages
    if not stages:
        stages = sorted(VALID_PIPELINE_STAGES)

    # Collect leads across requested stages
    all_leads: list[EnrichedLead] = []
    seen_ids: set[int] = set()
    for stage in stages:
        leads_in_stage = db.get_all_leads(
            stage=stage, min_score=body.min_score, limit=5000
        )
        for lead in leads_in_stage:
            if lead.id not in seen_ids:
                seen_ids.add(lead.id)  # type: ignore[arg-type]
                all_leads.append(lead)

    # Sort by score descending
    all_leads.sort(key=lambda l: l.score, reverse=True)

    xlsx_bytes = _build_xlsx(all_leads)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=salesleads_export.xlsx"},
    )
